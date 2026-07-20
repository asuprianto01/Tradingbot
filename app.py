"""
Web UI (Streamlit) for the IDX screeners.

Three read-only tabs over the tested logic modules:
  * Teknikal        -> screener.py / strategy.py / risk.py (signals + trade plan)
  * Quality Value   -> qvscreener.py / fundamentals.py     (fundamental scoring)
  * Mean Reversion  -> mean_reversion.py                   (daily swing setups)

Run it with:
    streamlit run app.py

It opens in your browser. Nothing here places an order.
"""

import io
import json
import os
import time
from concurrent.futures import ThreadPoolExecutor, as_completed

import pandas as pd
import streamlit as st

try:
    from dotenv import load_dotenv
    load_dotenv()          # baca TELEGRAM_BOT_TOKEN dari .env bila ada
except ImportError:
    pass

import fundamentals
import idx_data
import mean_reversion
import notify_telegram
import qvscreener
import screener
import strategy
import tracker
import universes

st.set_page_config(page_title="CuanRadar", page_icon="📡", layout="wide")

MAX_WORKERS = 8          # parallel fetches — indexes/ALL are slow one-by-one
QV_TOP_N = 40            # cap on how many top names to show for big universes
AUTO_CHECK_SEC = 60      # how often the auto-scan fragment wakes to check

# ATR multiplier stop loss per gaya — Long Term pakai 1W ATR yang sudah besar,
# jadi multiplier lebih kecil supaya stop tidak terlalu lebar.
STYLE_ATR_MULT = {
    "Swing Trading": 1.5,   # 1D ATR × 1.5 — standar swing
    "Intraday":      1.5,   # 1H ATR × 1.5 — ATR kecil, oke
    "Long Term":     1.0,   # 1W ATR × 1.0 — sudah inherently besar
}

# ATR multiplier untuk lebar zona entry (batas bawah zona beli).
# Target: zona bawah sekitar 1–3% di bawah harga pasar untuk semua gaya.
STYLE_ENTRY_ZONE_MULT = {
    "Swing Trading": 0.5,   # 1D ATR × 0.5 → ~1–2% di bawah — sudah pas
    "Intraday":      1.0,   # 1H ATR × 1.0 → ~0.5–1% — perlu lebih lebar agar bermakna
    "Long Term":     0.3,   # 1W ATR × 0.3 → ~1.5–3% — 1W ATR besar, perlu dipersempit
}

SIGNAL_COLORS = {
    strategy.SIGNAL_STRONG_BUY: "#0f9d58",
    strategy.SIGNAL_BUY: "#4caf50",
    strategy.SIGNAL_NEUTRAL: "#9e9e9e",
    strategy.SIGNAL_SELL: "#ef5350",
    strategy.SIGNAL_STRONG_SELL: "#c62828",
    mean_reversion.SIGNAL_WATCH: "#f9a825",
    mean_reversion.SIGNAL_NO_TRADE: "#9e9e9e",
}
MR_STATUS_COLORS = {
    mean_reversion.STATUS_STRONG: "#0f9d58",
    mean_reversion.STATUS_VALID: "#4caf50",
    mean_reversion.STATUS_WEAK: "#f9a825",
    mean_reversion.STATUS_NONE: "#9e9e9e",
}
ACCUM_LABELS = {
    "ACCUMULATION": "🟢 Akumulasi",
    "NEUTRAL": "⚪ Netral",
    "DISTRIBUTION": "🔴 Distribusi",
}
DIVERG_LABELS = {
    "BULLISH": "🟢 Bullish",
    "BEARISH": "🔴 Bearish",
    "NONE": "—",
}


VERDICT_COLORS = {
    "Quality Value - menarik": "#0f9d58",
    "Bagus tapi mahal": "#f9a825",
    "Biasa saja": "#9e9e9e",
    "Murah tapi berisiko": "#ef6c00",
    "Hindari": "#c62828",
    "Data kurang": "#9e9e9e",
}


LAST_SCAN_FILE = "last_scan.json"


def save_last_scan(tech):
    """Simpan hasil scan terakhir ke disk supaya bertahan lintas restart."""
    try:
        out = {"universe": tech.get("universe"), "styles": {}}
        for sname, ps in tech.get("per_style", {}).items():
            out["styles"][sname] = {
                "rows": ps["rows"], "recos": ps["recos"],
                "actionable": list(ps["actionable"]),
                "scanned_at": ps["scanned_at"],
                "errors": tech.get("errors_by_style", {}).get(sname, []),
                "analyses": {tk: a.to_dict() for tk, a in ps["analyses"].items()},
            }
        with open(LAST_SCAN_FILE, "w", encoding="utf-8") as f:
            json.dump(out, f)
    except (OSError, TypeError, ValueError):
        pass


def load_last_scan(current_universe):
    """Pulihkan hasil scan terakhir (hanya bila universe-nya sama)."""
    try:
        with open(LAST_SCAN_FILE, encoding="utf-8") as f:
            d = json.load(f)
    except (FileNotFoundError, json.JSONDecodeError, OSError):
        return None
    if d.get("universe") != current_universe:
        return None
    per_style, errors_by_style = {}, {}
    for sname, s in d.get("styles", {}).items():
        per_style[sname] = {
            "rows": s["rows"], "recos": s["recos"],
            "actionable": set(s.get("actionable", [])),
            "scanned_at": s.get("scanned_at", 0),
            "analyses": {tk: strategy.Analysis.from_dict(a)
                         for tk, a in s.get("analyses", {}).items()},
        }
        errors_by_style[sname] = s.get("errors", [])
    return {"per_style": per_style, "errors_by_style": errors_by_style,
            "universe": d.get("universe")}


def format_event_message(ev):
    """Pesan Telegram untuk sentuhan SL/TP (dengan trailing/breakeven)."""
    tkr, sty, gain = ev["ticker"], ev["style"], ev["gain_pct"]
    if ev["event"] == "SL":
        if gain > 0.1:
            tail = f"Profit terkunci {gain:+.1f}% (stop sudah di atas entry)."
        elif gain >= -0.1:
            tail = "Ditutup di BREAKEVEN (impas) — profit TP1 sudah aman."
        else:
            tail = f"Ditutup — LOSS {gain:+.1f}%."
        return f"🔴 STOP — {tkr} ({sty})\nHarga menyentuh stop {ev['level']:,}\n{tail}"

    # Take profit
    if ev.get("closed"):
        return (f"🎯 {ev['event']} (target tertinggi) — {tkr} ({sty})\n"
                f"Harga capai {ev['level']:,} · {gain:+.1f}% · Posisi selesai (WIN).")
    lines = [f"🟢 {ev['event']} tercapai — {tkr} ({sty})",
             f"Harga capai {ev['level']:,} · {gain:+.1f}%"]
    if ev.get("new_stop"):
        is_be = abs(ev["new_stop"] - ev["entry"]) < 1
        target = "breakeven (impas)" if is_be else f"{ev['new_stop']:,}"
        lines.append(f"🟡 Stop digeser ke {target} — profit diamankan.")
    return "\n".join(lines)


def format_entry_reminder(pos, current_price):
    """Pesan Telegram reminder: sinyal hari lalu masih di zona entry."""
    opened_date = pos.get("opened_at", "")[:10]
    ticker = pos["ticker"]
    style = pos["style"]
    entry_low = pos.get("entry_low", 0)
    entry_high = pos.get("entry_high", pos.get("entry_ref", 0))
    stop = pos.get("stop")
    tp1 = pos.get("tp1")
    gain_tp1 = (tp1 - current_price) / current_price * 100 if tp1 and current_price else None
    lines = [
        f"📌 Reminder — {ticker} ({style})",
        f"Sinyal BUY tgl {opened_date} masih relevan",
        f"📍 Harga {current_price:,.0f} — masih di zona entry ({entry_low:,.0f}–{entry_high:,.0f})",
    ]
    if stop:
        lines.append(f"🛑 SL: {stop:,.0f}")
    if tp1:
        tp_note = f" (+{gain_tp1:.1f}%)" if gain_tp1 is not None else ""
        lines.append(f"🎯 TP1: {tp1:,.0f}{tp_note}")
    lines.append("⚠️ Bukan nasihat keuangan.")
    return "\n".join(lines)


def color_signal(val):
    return f"color: {SIGNAL_COLORS.get(val, '#000')}; font-weight: 600"


def color_mr_status(val):
    return f"color: {MR_STATUS_COLORS.get(val, '#000')}; font-weight: 600"


def color_verdict(val):
    return f"color: {VERDICT_COLORS.get(val, '#000')}; font-weight: 600"


@st.cache_data(ttl=1800, show_spinner=False)
def company_profile(ticker):
    try:
        return fundamentals.get_profile(ticker)
    except Exception:
        return {"ticker": ticker, "name": None, "sector": None,
                "industry": None, "summary": None}


def show_company_header(ticker, prof=None):
    prof = prof or company_profile(ticker)
    st.markdown(f"### {ticker} — {prof.get('name') or 'nama tidak tersedia'}")
    meta = " · ".join(x for x in [prof.get("sector"), prof.get("industry")] if x)
    if meta:
        st.caption(f"🏢 {meta}")
    if prof.get("summary"):
        with st.expander("Profil & bidang usaha perusahaan"):
            st.write(prof["summary"])


def selected_ticker(event, frame):
    """Map a dataframe selection event to the 'Saham' value of the clicked row."""
    try:
        rows = event.selection.rows
    except Exception:
        rows = []
    if rows and 0 <= rows[0] < len(frame):
        return frame.iloc[rows[0]]["Saham"]
    return None


def summarize_factors(factors):
    """Split confluence factors into supporters (positive contribution) and
    detractors (negative), each sorted by magnitude of score×weight."""
    contribs = [(name, score * weight) for name, score, weight, _ in factors]
    pos = sorted([c for c in contribs if c[1] > 0.05], key=lambda x: -x[1])
    neg = sorted([c for c in contribs if c[1] < -0.05], key=lambda x: x[1])
    return pos, neg


# Tren & regime = PRASYARAT (sudah tampil di metrik "Tren"), bukan "alasan utama"
# untuk beli SEKARANG. Alasan utama = pemicu aktual (momentum/struktur/akumulasi).
_CONTEXT_FACTORS = {"HTF trend", "Market Regime", "Volatility", "ADX"}


def primary_reason(a):
    """Satu kalimat 'alasan utama' — keterangan faktor pemicu terkuat (bukan angka)."""
    if getattr(a, "smart_money_dip", False):
        return ("🎯 **Alasan utama — Smart-Money Dip (STRONG):** tren UP + StochRSI cross-up "
                "dari oversold + smart-money mengakumulasi (OBV naik, CMF positif).")
    if getattr(a, "srsi_dip", False) and a.signal in screener.BUY_SIGNALS:
        return ("💡 **Alasan utama:** tren UP + StochRSI cross-up dari oversold "
                "(akumulasi belum terkonfirmasi → BUY, bukan STRONG_BUY).")
    notes = {n: note for n, s, w, note in a.factors}
    pos, neg = summarize_factors(a.factors)
    if a.signal in screener.BUY_SIGNALS:
        drivers = [n for n, _ in pos if n not in _CONTEXT_FACTORS]
        if drivers:
            return f"💡 **Alasan utama:** {notes.get(drivers[0], drivers[0])} — didukung tren {a.trend}."
        return f"💡 **Alasan utama:** tren {a.trend} kuat & konfluensi indikator condong beli."
    if a.signal in (strategy.SIGNAL_SELL, strategy.SIGNAL_STRONG_SELL):
        det = [n for n, _ in neg if n not in _CONTEXT_FACTORS]
        head = notes.get(det[0], "tekanan jual dominan") if det else "tekanan jual dominan"
        return f"💡 **Alasan utama:** {head} — tren {a.trend}."
    return f"💡 **Alasan utama:** belum ada konfluensi jelas (tren {a.trend}). Tunggu konfirmasi."


def metric_table(fund, scores, specs):
    rows = []
    for key, spec in specs.items():
        v = fund.get(key)
        s = scores.get(key)
        rows.append({
            "Metrik": spec["label"],
            "Nilai": "-" if v is None else f"{v:,.2f}",
            "Skor (0-100)": "-" if s is None else f"{s:.0f}",
        })
    return pd.DataFrame(rows)


def format_signals_message(per_style, pairs=None, is_new=False):
    """
    Susun pesan Telegram dari hasil per-gaya.
    per_style: {gaya: {"analyses": {kode: Analysis}, "actionable": set()}}.
    pairs: iterable (gaya, kode) yang disertakan; None = semua yang actionable.
    """
    header_time = ""
    if hasattr(idx_data, "market_status"):
        header_time = idx_data.market_status()["wib"].strftime("%d %b %Y %H:%M") + " WIB"
    title = "📊 Sinyal BARU IDX" if is_new else "📊 Sinyal Screener IDX"
    lines = [f"{title} — {header_time}".rstrip(), f"Universe: {universe}", ""]

    by_style = {}
    if pairs is None:
        for sname, ps in per_style.items():
            for tkr in ps["actionable"]:
                by_style.setdefault(sname, []).append(tkr)
    else:
        for sname, tkr in pairs:
            by_style.setdefault(sname, []).append(tkr)

    any_item = False
    for sname in per_style:                       # urutan stabil sesuai preset
        tickers = by_style.get(sname)
        if not tickers:
            continue
        e, t = TRADING_STYLES[sname]
        analyses = per_style[sname]["analyses"]
        items = sorted((analyses[tk] for tk in tickers if tk in analyses),
                       key=lambda a: (screener.SIGNAL_RANK.get(a.signal, 2), a.score),
                       reverse=True)
        header_added = False
        for a in items:
            reco = screener.build_recommendation(a, cfg_for_style(sname))
            if not reco:
                continue
            if not header_added:
                lines.append(f"━ {sname} ({e}/{t}) ━")
                header_added = True
            any_item = True
            mark = "🟢🟢" if a.signal == strategy.SIGNAL_STRONG_BUY else "🟢"
            lines.append(f"{mark} {a.instrument} — {a.signal} (skor {a.score:+.2f})")
            lines.append(f"Zona beli {reco['entry_low']:,.0f}–{reco['entry_high']:,.0f} · SL {reco['stop']:,.0f}")
            lines.append(" · ".join(f"TP{i+1} {tg['price']:,.0f} (+{tg['gain_pct']:.1f}%)"
                                    for i, tg in enumerate(reco["targets"])))
            lines.append(f"Akum: {getattr(a, 'accumulation', 'NEUTRAL').title()}")
            lines.append("")

    if not any_item:
        lines.append("Tidak ada sinyal beli yang layak (reward:risk) saat ini.")
    else:
        lines.append("⚠️ Bukan nasihat keuangan. Data delayed ~15-20 mnt.")
    return "\n".join(lines)


def technical_detail(a, style_signals=None):
    show_company_header(a.instrument)
    cols = st.columns(4)
    cols[0].metric("Sinyal", a.signal)
    cols[1].metric("Skor confluence", f"{a.score:+.2f}")
    cols[2].metric("Tren (HTF)", a.trend)
    cols[3].metric("Akumulasi",
                   ACCUM_LABELS.get(getattr(a, "accumulation", "NEUTRAL"), "⚪ Netral"))

    _to = getattr(a, "turnover", None)
    if _to is not None:
        _min = cfg.get("min_turnover", 0)
        if _min and _to < _min:
            st.warning(f"💧 Likuiditas rendah — transaksi ~Rp {_to/1e9:.1f} M/hari "
                       f"(< ambang Rp {_min/1e9:.0f} M). Sinyal **tidak dijadikan rencana beli** "
                       "(slippage/ARB bikin backtest fiktif).")
        else:
            st.caption(f"💧 Likuiditas: transaksi ~Rp {_to/1e9:.1f} M/hari")

    # Alasan utama — satu kalimat menonjol
    st.info(primary_reason(a))
    if getattr(a, "srsi_dip", False) and a.signal not in screener.BUY_SIGNALS:
        st.caption("StochRSI: cross-up dari oversold terdeteksi, tapi sinyal butuh **tren UP** "
                   "(STRONG_BUY butuh akumulasi juga).")

    # Perbandingan lintas gaya (jika 'pantau semua gaya' aktif)
    if style_signals:
        chips = " · ".join(
            f"{sn} ({TRADING_STYLES[sn][0]}/{TRADING_STYLES[sn][1]}): "
            f"{sig} [{tr}]" for sn, (sig, tr) in style_signals.items())
        st.markdown(f"**Lintas gaya:** {chips}")

    # Pendukung/penghambat — ringkas (nama saja); rincian penuh dilipat di expander
    pos, neg = summarize_factors(a.factors)
    bits = []
    if pos:
        bits.append("✅ Pendukung: " + ", ".join(n for n, _ in pos[:3]))
    if neg:
        bits.append("⚠️ Penghambat: " + ", ".join(n for n, _ in neg[:3]))
    if bits:
        st.caption(" · ".join(bits))

    with st.expander("🔍 Rincian semua faktor (kenapa sinyal ini muncul)"):
        frows = [{
            "Faktor": n,
            "Arah": "Dukung ⬆" if s > 0.05 else ("Hambat ⬇" if s < -0.05 else "Netral"),
            "Skor": round(s, 2), "Bobot": w, "Keterangan": note,
        } for n, s, w, note in a.factors]
        st.dataframe(pd.DataFrame(frows), hide_index=True, use_container_width=True)

    reco = screener.build_recommendation(a, cfg_for_style(style_name))
    if reco:
        tps = reco.get("targets", [])
        ladder = " · ".join(f"TP{i+1} {t['price']:,.0f} (+{t['gain_pct']:.1f}%)"
                            for i, t in enumerate(tps))
        st.success(f"**Rencana beli** — Zona beli {reco['entry_low']:,.0f}–{reco['entry_high']:,.0f} "
                   f"(pasar {reco['entry']:,.0f}) · Stop Loss {reco['stop']:,.0f} · {ladder}")
        st.caption("Zona beli: idealnya akumulasi di dekat batas bawah (dekat support / pullback) "
                   "untuk reward:risk lebih baik. R:R & lot dihitung untuk entry di harga pasar (konservatif).")
    elif a.signal in screener.BUY_SIGNALS:
        st.info("Sinyal BELI tapi reward:risk belum layak — tunggu setup lebih baik.")
    else:
        st.info("Bukan setup beli saat ini (sinyal bukan BUY).")


def qv_detail(r):
    f = r["fund"]
    show_company_header(f["ticker"], {
        "name": f.get("name"), "sector": f.get("sector"),
        "industry": f.get("industry"), "summary": f.get("summary")})
    cols = st.columns(4)
    cols[0].metric("Verdict", r["verdict"])
    cols[1].metric("QV", f"{r['qv']:.0f}" if r["qv"] is not None else "-")
    cols[2].metric("Quality", f"{r['quality']:.0f}" if r["quality"] is not None else "-")
    cols[3].metric("Value", f"{r['value']:.0f}" if r["value"] is not None else "-")

    q, v = r["quality"], r["value"]
    if q is not None and v is not None:
        qtxt = "tinggi" if q >= 60 else ("sedang" if q >= 40 else "rendah")
        vtxt = "murah" if v >= 60 else ("wajar" if v >= 40 else "mahal")
        st.markdown(f"**Kesimpulan:** kualitas bisnis **{qtxt}** ({q:.0f}/100), "
                    f"valuasi **{vtxt}** ({v:.0f}/100) → **{r['verdict']}**.")

    col1, col2 = st.columns(2)
    with col1:
        st.markdown("**Value — kenapa murah/mahal:**")
        st.dataframe(metric_table(f, r["value_scores"], qvscreener.VALUE_METRICS),
                     hide_index=True, use_container_width=True)
    with col2:
        st.markdown("**Quality — kenapa bagus/lemah:**")
        st.dataframe(metric_table(f, r["quality_scores"], qvscreener.QUALITY_METRICS),
                     hide_index=True, use_container_width=True)

    if f.get("f_score") is not None:
        na = f.get("f_score_na", 0)
        na_note = f" · {na} kriteria n/a" if na else ""
        st.markdown(f"**Piotroski F-Score: {f['f_score']}/9**{na_note}")
        crit = [{"Kriteria": lbl,
                 "Nilai": "✅" if aw is True else ("❌" if aw is False else "n/a")}
                for lbl, aw in f.get("f_criteria", [])]
        if crit:
            st.dataframe(pd.DataFrame(crit), hide_index=True, use_container_width=True)

    if f.get("fair_value"):
        st.markdown(f"**Margin of Safety:** harga wajar (Graham) "
                    f"{f['fair_value']:,.0f} vs harga {f['price']:,.0f} → "
                    f"**{f['margin_of_safety']:+.1f}%**"
                    + (" (di bawah wajar 👍)" if (f['margin_of_safety'] or 0) > 0
                       else " (di atas wajar)"))


def scan_map(scan_tickers, fn, label):
    """Run fn(ticker) across all tickers in parallel, with a progress bar.
    Returns (results_by_ticker, errors)."""
    results, errors = {}, []
    total = len(scan_tickers)
    prog = st.progress(0.0, text=f"{label} 0/{total}")
    done = 0
    with ThreadPoolExecutor(max_workers=MAX_WORKERS) as ex:
        futures = {ex.submit(fn, t): t for t in scan_tickers}
        for fut in as_completed(futures):
            t = futures[fut]
            done += 1
            prog.progress(done / total, text=f"{label} {done}/{total}")
            try:
                results[t] = fut.result()
            except Exception as e:
                errors.append((t, str(e)))
    prog.empty()
    return results, errors


def cfg_for_style(style):
    """Config dengan ATR multiplier yang tepat per gaya (stop & zona entry)."""
    c = dict(cfg)
    c["risk"] = dict(cfg.get("risk", {}))
    c["risk"]["stop_atr_multiple"]    = STYLE_ATR_MULT.get(style, 1.5)
    c["risk"]["entry_zone_atr_mult"]  = STYLE_ENTRY_ZONE_MULT.get(style, 0.5)
    return c


def build_tech_row(a, row_cfg=None):
    """Dari sebuah Analysis: (row tabel, row ukuran-posisi | None, apakah actionable)."""
    reco = screener.build_recommendation(a, row_cfg if row_cfg is not None else cfg)
    tps = reco.get("targets") if reco else []
    if reco and not tps and reco.get("target"):
        tps = [{"price": reco["target"],
                "gain_pct": (reco["target"] - reco["entry"]) / reco["entry"] * 100,
                "rr": reco.get("reward_risk")}]

    def tp(i, key):
        return tps[i][key] if reco and len(tps) > i else None

    row = {
        "Saham": a.instrument, "Harga": a.price, "Tren": a.trend, "Sinyal": a.signal,
        "Akumulasi": ACCUM_LABELS.get(getattr(a, "accumulation", "NEUTRAL"), "⚪ Netral"),
        "Divergensi": DIVERG_LABELS.get(getattr(a, "divergence", "NONE"), "—"),
        "Transaksi (M/hr)": round((getattr(a, "turnover", None) or 0) / 1e9, 1),
        "Skor": round(a.score, 2),
        "Zona Beli": f"{reco['entry_low']:,.0f}–{reco['entry_high']:,.0f}" if reco else None,
        "Stop Loss": reco["stop"] if reco else None,
        "TP1": tp(0, "price"), "TP2": tp(1, "price"), "TP3": tp(2, "price"),
        "R:R (TP1)": round(reco["reward_risk"], 1) if reco else None,
        "Potensi% (TP akhir)": round(tps[-1]["gain_pct"], 1) if reco and tps else None,
        "Support": a.nearest_support, "Resistance": a.nearest_resistance,
        "Catatan": ("" if reco else ("⚠️ R:R tak layak (resistance dekat)"
                    if a.signal in screener.BUY_SIGNALS else "")),
        "_rank": screener.SIGNAL_RANK.get(a.signal, 2),
    }
    reco_row = None
    if reco:
        reco_row = {"Saham": a.instrument, "Sinyal": reco["signal"],
                    "Entry": reco["entry"], "Stop Loss": reco["stop"],
                    "TP1": reco["target"], "R:R": round(reco["reward_risk"], 1),
                    "Lot": reco["lots"], "Modal (Rp)": round(reco["cost"]),
                    "Risiko (Rp)": round(reco["risk_amount"])}
    return row, reco_row, (reco is not None)


def mr_primary_reason(a):
    if a.signal == mean_reversion.SIGNAL_BUY and a.reasons:
        head = " + ".join(a.reasons[:2])
        return f"Setup mean reversion aktif karena {head.lower()}"
    if a.signal == mean_reversion.SIGNAL_WATCH and a.reasons:
        return f"Watchlist dulu: {a.reasons[0]}"
    if a.blockers:
        return f"Belum layak entry: {a.blockers[0]}"
    return "Belum ada kombinasi rule yang cukup kuat."


def build_mr_row(a):
    plan = mean_reversion.build_position_plan(
        a, capital=capital, risk_pct=risk_pct, lot_size=cfg.get("lot_size", 100)
    )
    row = {
        "Saham": a.instrument,
        "Sinyal": a.signal,
        "Status": a.status,
        "Skor": a.score,
        "Harga": a.price,
        "Chg%": round(a.change_pct, 1) if a.change_pct is not None else None,
        "RSI 14": round(a.rsi14, 1) if a.rsi14 is not None else None,
        "Jarak ke MA20%": round(a.close_vs_ma20_pct, 1) if a.close_vs_ma20_pct is not None else None,
        "Vol Ratio": round(a.volume_ratio, 2) if a.volume_ratio is not None else None,
        "Entry": a.entry_price,
        "Stop Loss": a.stop_loss,
        "Target 1": a.target_1,
        "Target 2": a.target_2,
        "R:R": round(a.risk_reward, 2) if a.risk_reward is not None else None,
        "Support": a.nearest_support,
        "Resistance": a.nearest_resistance,
        "Alasan": a.reasons[0] if a.reasons else (a.blockers[0] if a.blockers else ""),
        "_rank": 2 if a.signal == mean_reversion.SIGNAL_BUY else (
            1 if a.signal == mean_reversion.SIGNAL_WATCH else 0
        ),
    }
    reco_row = None
    if plan:
        reco_row = {
            "Saham": a.instrument,
            "Sinyal": a.signal,
            "Entry": plan["entry"],
            "Stop Loss": plan["stop"],
            "Target 1": plan["target_1"],
            "Target 2": plan["target_2"],
            "R:R": round(plan["reward_risk"], 2) if plan["reward_risk"] is not None else None,
            "Lot": plan["lots"],
            "Modal (Rp)": round(plan["cost"]),
            "Risiko (Rp)": round(plan["risk_amount"]),
        }
    return row, reco_row


def _journal_mean_reversion_scan(analyses, mode_label):
    style = f"Mean Reversion {mode_label}"
    positions = tracker.load_open()
    cooldowns = tracker.load_cooldowns()
    bars = {tk: (a.high, a.low) for tk, a in analyses.items()}
    hit_events = tracker.update_positions(positions, bars, style=style, cooldowns=cooldowns)
    new_opens = []
    for tk, a in analyses.items():
        reco = mean_reversion.build_tracker_recommendation(a)
        if reco and tracker.open_position(
                positions, style, tk, a.signal, reco,
                current_low=a.low, cooldowns=cooldowns):
            new_opens.append(tk)
    tracker.save_open(positions)
    tracker.save_cooldowns(cooldowns)
    return hit_events, new_opens


def mean_reversion_detail(a):
    show_company_header(a.instrument)
    cols = st.columns(4)
    cols[0].metric("Sinyal", a.signal)
    cols[1].metric("Status", a.status)
    cols[2].metric("Skor", f"{a.score}/100")
    cols[3].metric("Mode", a.mode.title())

    st.info(mr_primary_reason(a))
    if a.crash_day:
        st.warning("Harga turun terlalu dalam dalam satu hari, jadi setup disaring oleh rule PRD.")

    c1, c2, c3, c4 = st.columns(4)
    c1.metric("RSI 14", f"{a.rsi14:.1f}" if a.rsi14 is not None else "-")
    c2.metric("Volume Ratio", f"{a.volume_ratio:.2f}x" if a.volume_ratio is not None else "-")
    c3.metric("Support Gap", f"{a.support_gap_pct:.1f}%" if a.support_gap_pct is not None else "-")
    c4.metric("R:R", f"{a.risk_reward:.2f}:1" if a.risk_reward is not None else "-")

    if a.reasons:
        st.caption("Pendukung: " + " | ".join(a.reasons))
    if a.blockers:
        st.caption("Penghambat: " + " | ".join(a.blockers))

    checklist = [
        {"Rule": "Touch lower Bollinger", "Hasil": "Lolos" if a.touched_lower_band else "Belum", "Catatan": f"Lower BB {a.bb_lower:,.0f}" if a.bb_lower is not None else "-"},
        {"Rule": "RSI oversold", "Hasil": "Lolos" if a.rsi14 is not None and ((a.mode == mean_reversion.MODE_CONSERVATIVE and a.rsi14 <= 35) or (a.mode == mean_reversion.MODE_AGGRESSIVE and a.rsi14 <= 30)) else "Belum", "Catatan": f"RSI {a.rsi14:.1f}" if a.rsi14 is not None else "-"},
        {"Rule": "Trend filter", "Hasil": "Lolos" if a.trend_ok else "Belum", "Catatan": f"MA50 {a.ma50:,.0f} / MA200 {a.ma200:,.0f}" if a.ma50 is not None and a.ma200 is not None else "-"},
        {"Rule": "Near support", "Hasil": "Lolos" if a.near_support else "Belum", "Catatan": f"Support {a.nearest_support:,.0f}" if a.nearest_support is not None else "-"},
        {"Rule": "Volume >= 0.8x", "Hasil": "Lolos" if a.volume_ratio is not None and a.volume_ratio >= 0.8 else "Belum", "Catatan": f"{a.volume_ratio:.2f}x" if a.volume_ratio is not None else "-"},
        {"Rule": "Rebound candle", "Hasil": "Lolos" if a.rebound else "Belum", "Catatan": "Close kembali kuat dari low/band"},
        {"Rule": "Risk/reward minimum", "Hasil": "Lolos" if a.risk_reward is not None and ((a.mode == mean_reversion.MODE_CONSERVATIVE and a.risk_reward >= 1.5) or (a.mode == mean_reversion.MODE_AGGRESSIVE and a.risk_reward >= 1.2)) else "Belum", "Catatan": f"{a.risk_reward:.2f}:1" if a.risk_reward is not None else "-"},
    ]
    with st.expander("🔍 Checklist rule PRD"):
        st.dataframe(pd.DataFrame(checklist), hide_index=True, use_container_width=True)

    plan = mean_reversion.build_position_plan(
        a, capital=capital, risk_pct=risk_pct, lot_size=cfg.get("lot_size", 100)
    )
    if plan:
        st.success(
            f"**Rencana beli** — Entry {plan['entry']:,.0f} · Stop Loss {plan['stop']:,.0f} · "
            f"Target 1 {plan['target_1']:,.0f} · Target 2 {plan['target_2']:,.0f} · "
            f"R:R {plan['reward_risk']:.2f}:1"
        )
        st.caption(
            f"Ukuran posisi mengikuti sidebar: {plan['lots']} lot (~Rp {plan['cost']:,.0f}) "
            f"dengan risiko ~Rp {plan['risk_amount']:,.0f}."
        )
    elif a.signal == mean_reversion.SIGNAL_WATCH:
        st.info("Setup mulai menarik, tapi masih butuh konfirmasi tambahan sebelum dianggap entry-ready.")
    else:
        st.info("Belum ada setup mean reversion yang layak dieksekusi saat ini.")

    st.caption("Data Yahoo Finance delayed ~15-20 menit. Bukan nasihat keuangan.")


def render_mr_backtest(mode, mode_label):
    st.divider()
    st.subheader("Backtest Mean Reversion")
    b1, b2, b3 = st.columns([1, 1, 1.4])
    bt_period = b1.selectbox("Periode", ["2y", "5y", "10y"], index=1, key="mr_bt_period")
    bt_horizon = b2.number_input("Max holding (hari)", min_value=5, max_value=60, value=20, step=5, key="mr_bt_horizon")
    run_bt = b3.button("▶️ Jalankan backtest", type="primary", key="mr_bt_run")
    b3.caption(f"Mode {mode_label} · modal Rp {capital:,.0f} · risiko {risk_pct}%/trade")

    if run_bt:
        if not scan_tickers:
            st.warning("Universe kosong.")
        else:
            bt_results, bt_errors = scan_map(
                scan_tickers,
                lambda tk: mean_reversion.backtest_one(
                    tk,
                    mode=mode,
                    period=bt_period,
                    horizon=int(bt_horizon),
                    initial_capital=capital,
                    risk_per_trade=risk_pct / 100.0,
                ),
                "Backtest Mean Reversion",
            )
            combined = mean_reversion.combine_backtests(
                bt_results.values(), initial_capital=capital, risk_per_trade=risk_pct / 100.0
            )
            st.session_state["mr_backtest"] = {
                "combined": combined,
                "errors": bt_errors,
                "sig": (universe, tuple(scan_tickers), mode, bt_period, int(bt_horizon), capital, risk_pct),
            }

    bt = st.session_state.get("mr_backtest")
    if not bt:
        return
    metrics = bt["combined"]["metrics"]
    m1, m2, m3, m4 = st.columns(4)
    m1.metric("Total trade", metrics["total_trades"])
    m2.metric("Win rate", f"{metrics['win_rate'] * 100:.1f}%")
    m3.metric("Expectancy", f"{metrics['expectancy_R']:+.2f}R")
    pf = "∞" if metrics["profit_factor"] == float("inf") else f"{metrics['profit_factor']:.2f}"
    m4.metric("Profit factor", pf)
    m5, m6, m7, m8 = st.columns(4)
    m5.metric("Return", f"{metrics['total_return_pct']:+.1f}%")
    m6.metric("Max drawdown", f"{metrics['max_drawdown_pct']:.1f}%")
    m7.metric("Avg holding", f"{metrics['avg_holding_bars']:.1f} hari")
    m8.metric("TP rate", f"{metrics['tp_rate'] * 100:.1f}%")

    curve = bt["combined"].get("equity_curve", [])
    if curve:
        st.line_chart(pd.DataFrame({"Equity": curve}))
    trades = pd.DataFrame(bt["combined"].get("trades", []))
    if trades.empty:
        st.info("Backtest tidak menemukan trade untuk parameter ini.")
    else:
        st.dataframe(
            trades.rename(columns={
                "ticker": "Saham", "signal_date": "Tanggal Sinyal", "exit_date": "Tanggal Exit",
                "entry": "Entry", "stop": "Stop", "target": "Target", "exit": "Exit",
                "outcome": "Hasil", "holding_bars": "Holding", "r_multiple": "R",
                "pnl": "PnL", "score": "Skor", "rsi14": "RSI 14", "risk_reward": "R:R",
            }).style.format({
                "Entry": "{:,.0f}", "Stop": "{:,.0f}", "Target": "{:,.0f}", "Exit": "{:,.0f}",
                "R": "{:+.2f}", "PnL": "{:,.0f}", "RSI 14": "{:.1f}", "R:R": "{:.2f}",
            }, na_rep="-"),
            use_container_width=True,
            hide_index=True,
        )
        csv_bt = trades.to_csv(index=False).encode("utf-8-sig")
        st.download_button(
            "⬇️ Export trade list backtest (.csv)",
            csv_bt,
            file_name=f"mean_reversion_backtest_{mode}_{bt_period}.csv",
            mime="text/csv",
            key="mr_bt_export_csv",
        )
    if bt["errors"]:
        with st.expander(f"⚠️ {len(bt['errors'])} saham gagal saat backtest"):
            for t, e in bt["errors"]:
                st.write(f"**{t}**: {e}")


def _show_tech_errors(data):
    errs = data.get("errors_by_style", {}).get(style_name, [])
    if errs:
        with st.expander(f"⚠️ {len(errs)} saham gagal diambil ({style_name})"):
            for t, e in errs:
                st.write(f"**{t}**: {e}")


# ---------------------------------------------------------------------------
# Sidebar — shared settings
# ---------------------------------------------------------------------------
cfg = screener.load_config("config_screener.json")

# Brand
st.sidebar.markdown("## 📡 CuanRadar")
st.sidebar.caption("Screener Saham IDX · Teknikal · Quality Value · Mean Reversion")
st.sidebar.divider()

# Universe & daftar saham
st.sidebar.subheader("🗂️ Universe")
_univ_options = universes.available()
universe = st.sidebar.selectbox("Daftar saham", _univ_options,
                                index=_univ_options.index("ALL"),
                                label_visibility="collapsed")

if "selected_text" not in st.session_state:
    st.session_state["selected_text"] = "\n".join(cfg["tickers"])


def _save_selected():
    tick = [t.strip().upper() for t in st.session_state["selected_text"].splitlines() if t.strip()]
    screener.save_tickers(tick)
    st.session_state["_saved"] = len(tick)


def _clear_selected():
    st.session_state["selected_text"] = ""


tickers_text = st.sidebar.text_area(
    "Daftar 'Selected'",
    key="selected_text", height=120,
    help="Dipakai saat Universe = Selected. Satu kode per baris, tanpa .JK. "
         "Klik Simpan agar jadi default permanen.",
    placeholder="BBCA\nBBRI\nTLKM",
)
bcol1, bcol2 = st.sidebar.columns(2)
bcol1.button("💾 Simpan", on_click=_save_selected, use_container_width=True,
             help="Simpan sebagai default permanen di config.json")
bcol2.button("🗑️ Kosongkan", on_click=_clear_selected, use_container_width=True)
if st.session_state.pop("_saved", None) is not None:
    st.sidebar.success("Daftar tersimpan.")

selected_list = [t.strip().upper() for t in tickers_text.splitlines() if t.strip()]

if universe == "FILTER":
    import refresh_universe
    _cache = refresh_universe.load()
    if _cache is None:
        st.sidebar.warning("Cache universe belum ada.\nJalankan: `python refresh_universe.py`")
        scan_tickers = []
    else:
        _c1, _c2 = st.sidebar.columns(2)
        _min_price = _c1.number_input("Harga min (Rp)", min_value=0,
                                      value=int(_cache["min_price"]), step=50)
        _min_mcap_t = _c2.number_input("Mcap min (T)", min_value=0.0,
                                       value=round(_cache["min_mcap"] / 1e12, 1),
                                       step=0.5, format="%.1f")
        scan_tickers = universes.filtered(_min_price, _min_mcap_t * 1e12)
        st.sidebar.caption(f"Cache: {_cache['count']} saham · {_cache['generated']}")
else:
    scan_tickers = universes.get_universe(universe, selected_list)

n_label = f"**{len(scan_tickers)}** saham · {universe}"
if len(scan_tickers) > 120:
    st.sidebar.warning(f"{n_label}  \nUniverse besar — scan lambat & rentan rate-limit Yahoo.")
else:
    st.sidebar.caption(n_label)

st.sidebar.divider()

# Gaya trading & rotasi (sering diubah → tetap terlihat)
st.sidebar.subheader("📈 Gaya Trading")

TRADING_STYLES = {
    "Swing Trading": ("1d", "1wk"),
    "Intraday": ("1h", "1d"),
    "Long Term": ("1wk", "1mo"),
}
STYLE_ROTATION = list(TRADING_STYLES)

# Auto-scan cadence per gaya (detik) — ikut durasi bar timeframe entry-nya.
# Scan lebih sering dari ini sia-sia: candle belum ganti & data Yahoo delay ~15 mnt.
# Hanya berlaku saat pasar buka.
STYLE_INTERVALS = {
    "Intraday":      15 * 60,     # bar 1H → cek tiap 15 menit
    "Swing Trading": 60 * 60,     # bar 1D → tiap 60 menit (yang penting jelang tutup)
    "Long Term":     24 * 3600,   # bar 1W → cukup 1× per hari
}

def _style_interval(s):
    return STYLE_INTERVALS.get(s, 60 * 60)

def _due_style(market_open, rotate_styles, displayed, per_style):
    """Gaya yang jatuh tempo untuk auto-scan (interval per-gaya, jam bursa saja).
    Prioritaskan gaya yang sedang ditampilkan; jika tidak ada yang tampil, ambil
    gaya paling telat. None bila belum ada yang jatuh tempo."""
    if not market_open:
        return None
    candidates = STYLE_ROTATION if rotate_styles else [displayed]
    now = time.time()
    due = []
    for s in candidates:
        last = per_style.get(s, {}).get("scanned_at", 0)
        overdue = (now - last) - _style_interval(s)
        if overdue >= 0:
            due.append((overdue, s))
    if not due:
        return None
    for _, s in due:
        if s == displayed:
            return displayed
    return max(due)[1]

rotate_styles = st.sidebar.checkbox(
    "🔁 Rotasi semua gaya", value=True,
    help="Auto-scan semua gaya, masing-masing dengan interval sendiri "
         "(Intraday 15 mnt · Swing 60 mnt · Long Term harian). "
         "Matikan untuk auto-scan hanya gaya yang dipilih.")
style_name = st.sidebar.selectbox(
    "Gaya ditampilkan", list(TRADING_STYLES), index=0,
    help="Swing = 1D/1W (disarankan) · Intraday = 1H/1D · Long Term = 1W/1M.")
entry_tf, trend_tf = TRADING_STYLES[style_name]
st.sidebar.caption(f"TF: entry **{entry_tf}** / tren **{trend_tf}**")

st.sidebar.divider()

# Pengaturan lanjutan — jarang diubah, disimpan dalam expander
with st.sidebar.expander("⚙️ Pengaturan lanjutan"):
    min_signal = st.selectbox(
        "Min sinyal (mode Selected)",
        [strategy.SIGNAL_STRONG_SELL, strategy.SIGNAL_SELL, strategy.SIGNAL_NEUTRAL,
         strategy.SIGNAL_BUY, strategy.SIGNAL_STRONG_BUY],
        index=2,
        help="Filter sinyal minimum yang ditampilkan saat Universe = Selected.",
    )
    capital = st.number_input("Modal (Rp)", min_value=1_000_000,
                               value=int(cfg["capital"]), step=1_000_000)
    risk_pct = st.number_input("Risiko per trade (%)", min_value=0.1,
                               value=float(cfg["risk"].get("risk_per_trade_pct", 1.0)),
                               step=0.1)

    st.caption("Modal & risiko → hanya memengaruhi **ukuran lot**, bukan sinyal/SL/TP.")
    st.caption("↻ Interval auto-scan per gaya (jam bursa saja): "
               "**Intraday** 15 mnt · **Swing** 60 mnt · **Long Term** harian. "
               "Ikut durasi bar timeframe entry-nya — lebih sering = sia-sia (Yahoo delay ~15 mnt).")

cfg["capital"] = capital
cfg["risk"] = dict(cfg.get("risk", {}))
cfg["risk"]["risk_per_trade_pct"] = risk_pct

if st.sidebar.button("🧹 Reset hasil scan", use_container_width=True,
                     help="Hapus hasil scan dari memori (jurnal & posisi terbuka TIDAK terhapus)"):
    st.session_state.pop("tech", None)
    st.session_state.pop("qv", None)
    st.rerun()

# --- Telegram ---
with st.sidebar.expander("📤 Telegram"):
    _tg_token, _tg_recipients = notify_telegram.load_config()
    st.caption("🔑 Token: " + ("✅ terpasang" if _tg_token
                               else "❌ belum diset (.env TELEGRAM_BOT_TOKEN)"))
    auto_send = st.checkbox("Auto-kirim saat ada sinyal baru", value=True,
                            help="Otomatis kirim ke Telegram begitu muncul sinyal BELI baru "
                                 "yang actionable. Perlu token & penerima terisi.")
    _new_id = st.text_input("Tambah penerima (chat_id)", key="tg_new_id",
                            placeholder="mis. 123456789 atau -100xxx (grup)")
    if st.button("➕ Tambah", use_container_width=True, key="tg_add"):
        if _new_id.strip():
            notify_telegram.save_recipients(_tg_recipients + [_new_id.strip()])
            st.rerun()
    if _tg_recipients:
        st.caption(f"Penerima ({len(_tg_recipients)}): " + ", ".join(_tg_recipients))
        _rm = st.selectbox("Hapus penerima", ["—"] + _tg_recipients, key="tg_rm")
        if st.button("🗑️ Hapus penerima", use_container_width=True, key="tg_del") and _rm != "—":
            notify_telegram.save_recipients([r for r in _tg_recipients if r != _rm])
            st.rerun()
    else:
        st.caption("Belum ada penerima.")

st.sidebar.divider()
st.sidebar.caption("Data: Yahoo Finance (delayed ~15-20 mnt)  \nBukan nasihat keuangan.")

st.markdown(
    "<h1 style='margin-bottom:0'>📡 CuanRadar</h1>"
    "<p style='color:#666; margin-top:2px; font-size:0.95rem'>"
    "Screener Saham IDX &nbsp;·&nbsp; Teknikal &amp; Quality Value &nbsp;·&nbsp; "
    "Data Yahoo Finance delayed ~15–20 mnt &nbsp;·&nbsp; <em>Bukan nasihat keuangan</em></p>",
    unsafe_allow_html=True,
)


@st.cache_data(ttl=60, show_spinner=False)
def _quote_freshness():
    return idx_data.quote_freshness()


def render_freshness_bar():
    # Guard: if an older idx_data module is still loaded (before restart),
    # skip the bar instead of crashing the whole app.
    if not hasattr(idx_data, "market_status"):
        st.warning("🔄 Modul data baru belum termuat — **restart app** "
                   "(tutup terminal → buka run_screener_ui.bat) agar indikator "
                   "kesegaran & timeframe 1h aktif.")
        st.caption(f"📂 {universe} · {len(scan_tickers)} saham")
        return
    status = idx_data.market_status()          # cheap, always current
    fresh = _quote_freshness()                 # cached 60s (network)
    badge = "🟢 Pasar BUKA" if status["open"] else f"🔴 Pasar {status['label']}"
    parts = [badge, f"🕒 {status['wib']:%a %H:%M} WIB", f"📂 {universe} · {len(scan_tickers)} saham"]
    if fresh:
        if status["open"]:
            parts.append(f"📡 Kuotasi {fresh['quote_wib']:%H:%M} WIB · delay ~{fresh['delay_min']:.0f} mnt")
        else:
            parts.append(f"📡 Harga penutupan {fresh['quote_wib']:%d %b %H:%M} WIB")
    if status["open"]:
        st.info(" · ".join(parts) + "  \n⚠️ Pasar buka: candle timeframe berjalan (1h/1d) **belum final** — sinyal bisa berubah saat bar tutup.")
    else:
        st.success(" · ".join(parts))


render_freshness_bar()

tab_tech, tab_qv, tab_mr = st.tabs(["📈 Teknikal", "💎 Quality Value", "📉 Mean Reversion"])


# ---------------------------------------------------------------------------
# Tab 1 — technical screener
# ---------------------------------------------------------------------------
def render_technical():
    market_open = (idx_data.market_status()["open"]
                   if hasattr(idx_data, "market_status") else True)
    mcol1, mcol2 = st.columns([1, 3])
    manual = mcol1.button("🔍 Scan sekarang", type="primary", key="scan_now")
    if market_open:
        mcol2.caption("🟢 Pasar buka — auto-scan per gaya aktif (Intraday 15m · Swing 60m · "
                      "Long Term harian). Tombol = scan gaya ini sekarang.")
    else:
        mcol2.caption("🔴 Pasar tutup — auto-scan nonaktif (hemat request). "
                      "Klik **Scan sekarang** untuk scan manual gaya ini.")

    _sig = (universe, tuple(scan_tickers))
    _tech = st.session_state.get("tech")
    # Sesi baru: coba pulihkan hasil scan terakhir dari disk (biar tak kosong).
    if _tech is None:
        _tech = load_last_scan(universe)
        if _tech is not None:
            _tech["sig"] = _sig
            st.session_state["tech"] = _tech
    # Reset struktur bila universe/daftar berubah, atau belum ada.
    if (_tech is None) or (_tech.get("sig") != _sig) or ("per_style" not in _tech):
        _tech = {"per_style": {}, "errors_by_style": {}, "sig": _sig, "universe": universe}
        st.session_state["tech"] = _tech

    # Auto-scan per gaya (interval sendiri) HANYA saat pasar buka.
    # Tombol memaksa scan gaya yang ditampilkan kapan saja.
    if manual:
        _style = style_name
    else:
        _style = _due_style(market_open, rotate_styles, style_name, _tech["per_style"])

    if _style is not None:
        _e, _t = TRADING_STYLES[_style]
        res, errors = scan_map(
            scan_tickers, lambda tk: screener.screen_one(tk, _e, _t), f"Scan {_style}")
        _style_cfg = cfg_for_style(_style)
        ps = {"rows": [], "recos": [], "analyses": {}, "actionable": set(),
              "scanned_at": time.time()}
        for a in res.values():
            row, reco_row, ok = build_tech_row(a, _style_cfg)
            ps["rows"].append(row)
            ps["analyses"][a.instrument] = a
            if reco_row:
                ps["recos"].append(reco_row)
                ps["actionable"].add(a.instrument)
        _tech["per_style"][_style] = ps          # ps["scanned_at"] = waktu scan gaya ini
        _tech["errors_by_style"][_style] = errors

        # --- Tracker: jurnal + notifikasi SL/TP + buka sinyal baru ---
        _positions = tracker.load_open()
        _cooldowns = tracker.load_cooldowns()
        _bars = {tk: (getattr(ps["analyses"][tk], "high", ps["analyses"][tk].price),
                      getattr(ps["analyses"][tk], "low", ps["analyses"][tk].price))
                 for tk in ps["analyses"]}
        # Hanya cek posisi gaya yang baru discan (pakai timeframe-nya sendiri).
        _exit_cfg = cfg.get("exit", {})
        _trail = _exit_cfg.get("trail_r_mult", 1.5) if _exit_cfg.get("trailing", True) else None
        _hit_events = tracker.update_positions(
            _positions, _bars, style=_style, cooldowns=_cooldowns, trail_mult=_trail)

        _new_opens = []                       # (style, ticker) sinyal baru dibuka
        for tk in ps["actionable"]:
            a = ps["analyses"][tk]
            reco = screener.build_recommendation(a, _style_cfg)
            if reco and tracker.open_position(
                    _positions, _style, tk, a.signal, reco,
                    current_low=getattr(a, "low", None),
                    cooldowns=_cooldowns):
                _new_opens.append((_style, tk))

        # Reminder harian: posisi hari lalu yang harganya masih di zona entry (sekali sehari)
        _prices = {tk: ps["analyses"][tk].price for tk in ps["analyses"]}
        _reminders = tracker.get_entry_zone_reminders(_positions, _prices)
        if _reminders:
            tracker.mark_reminded(_positions, [k for k, _, _ in _reminders])

        tracker.save_open(_positions)
        tracker.save_cooldowns(_cooldowns)

        # Notifikasi Telegram (bila diaktifkan & terkonfigurasi)
        if auto_send:
            _token, _recips = notify_telegram.load_config()
            if _token and _recips:
                for _ev in _hit_events:
                    notify_telegram.broadcast(format_event_message(_ev), _token, _recips)
                if _new_opens:
                    _msg = format_signals_message(_tech["per_style"], pairs=_new_opens, is_new=True)
                    notify_telegram.broadcast(_msg, _token, _recips)
                for _, _rpos, _rprice in _reminders:
                    notify_telegram.broadcast(format_entry_reminder(_rpos, _rprice), _token, _recips)
        if _hit_events:
            st.toast(f"📈 {len(_hit_events)} level (SL/TP) tersentuh")
        if _new_opens:
            st.toast(f"📤 {len(_new_opens)} sinyal baru dibuka ({_style})")
        if _reminders:
            st.toast(f"📌 {len(_reminders)} sinyal kemarin masih di zona entry")

        save_last_scan(_tech)          # simpan hasil scan agar bertahan restart

    data = st.session_state.get("tech")
    if not data.get("per_style"):
        if market_open:
            st.info("Menyiapkan scan pertama…")
        else:
            st.info("🔴 Pasar tutup & belum ada data. Klik **Scan sekarang** untuk memuat.")
        render_journal()
        return

    # Status: waktu scan terakhir + scan berikutnya per interval-gaya
    _times = " · ".join(
        (f"{s} {time.strftime('%H:%M', time.localtime(data['per_style'][s]['scanned_at']))}"
         if s in data["per_style"] else f"{s} —")
        for s in STYLE_ROTATION)
    if market_open:
        _cands = STYLE_ROTATION if rotate_styles else [style_name]
        _now = time.time()
        _left, _next_style = min(
            (max(0, int(_style_interval(s) - (_now - data["per_style"].get(s, {}).get("scanned_at", 0)))), s)
            for s in _cands)
        _eta = f"~{_left // 60} mnt" if _left >= 60 else f"~{_left}s"
        _status = f"🔁 Scan berikutnya **{_next_style}** dalam {_eta}"
    else:
        _status = "🔴 Pasar tutup — auto-scan berhenti (klik Scan untuk manual)"
    st.caption(f"{_status} · terakhir: {_times} · "
               f"auto-kirim: {'aktif' if auto_send else 'nonaktif'}")
    counts = " · ".join(
        (f"**{s}**: {len(data['per_style'][s]['actionable'])}" if s in data["per_style"]
         else f"**{s}**: —") for s in STYLE_ROTATION)
    st.info(f"🎯 Sinyal actionable per gaya → {counts}")

    if style_name not in data["per_style"]:
        st.warning(f"Gaya '{style_name}' belum discan (menunggu giliran rotasi). "
                   "Pilih gaya lain yang sudah discan, atau tunggu.")
        render_journal()
        return
    ps = data["per_style"][style_name]
    if not ps["rows"]:
        st.error("Tidak ada data yang berhasil diambil.")
        _show_tech_errors(data)
        render_journal()
        return

    df = pd.DataFrame(ps["rows"]).sort_values(["_rank", "Skor"], ascending=False)
    is_selected = data["universe"] == "Selected"
    if is_selected:
        min_rank = screener.SIGNAL_RANK.get(min_signal, 2)
        visible = df[df["_rank"] >= min_rank]
    else:
        visible = df[df["Sinyal"].isin(screener.BUY_SIGNALS)]
    visible = visible.drop(columns=["_rank"])
    df = df.drop(columns=["_rank"])

    e, t = TRADING_STYLES[style_name]
    st.subheader(f"{style_name} — TF {e}/{t}")
    c1, c2, c3, c4 = st.columns(4)
    c1.metric("Saham discan", len(df))
    c2.metric("Sinyal BUY+", int(df["Sinyal"].isin(screener.BUY_SIGNALS).sum()))
    c3.metric("Ditampilkan", len(visible))
    c4.metric("Gagal ambil data", len(data.get("errors_by_style", {}).get(style_name, [])))

    st.caption("👆 Klik satu baris untuk melihat profil perusahaan & alasan sinyalnya.")
    if visible.empty:
        st.warning("Tidak ada saham yang lolos filter untuk gaya ini.")
    else:
        visible = visible.reset_index(drop=True)
        event = st.dataframe(
            visible.style.map(color_signal, subset=["Sinyal"]).format(
                {"Harga": "{:,.0f}", "Skor": "{:+.2f}",
                 "Stop Loss": "{:,.0f}", "TP1": "{:,.0f}", "TP2": "{:,.0f}",
                 "TP3": "{:,.0f}", "R:R (TP1)": "{:.1f}", "Potensi% (TP akhir)": "{:+.1f}",
                 "Support": "{:,.0f}", "Resistance": "{:,.0f}"}, na_rep="-"),
            use_container_width=True, hide_index=True,
            on_select="rerun", selection_mode="single-row", key="tech_table")
        picked = selected_ticker(event, visible)
        if picked and picked in ps["analyses"]:
            st.divider()
            style_sigs = {}
            for sn in data["per_style"]:
                aa = data["per_style"][sn]["analyses"].get(picked)
                if aa:
                    style_sigs[sn] = (aa.signal, aa.trend)
            technical_detail(ps["analyses"][picked],
                             style_sigs if len(style_sigs) > 1 else None)
    st.caption("TP1 = target utama (min. 2R / resistance terdekat) · "
               "TP2–TP3 = resistance (swing high) berikutnya = potensi lanjutan.")

    st.subheader("💡 Ukuran Posisi (lot)")
    st.caption(f"Modal Rp {capital:,.0f} · risiko {risk_pct}%/trade · long-only · gaya {style_name}")
    if not ps["recos"]:
        st.write("Tidak ada setup BELI yang memenuhi syarat reward:risk untuk gaya ini.")
    else:
        st.dataframe(
            pd.DataFrame(ps["recos"]).style.map(color_signal, subset=["Sinyal"]).format(
                {"Entry": "{:,.0f}", "Stop Loss": "{:,.0f}", "TP1": "{:,.0f}",
                 "R:R": "{:.1f}", "Modal (Rp)": "{:,.0f}", "Risiko (Rp)": "{:,.0f}"}),
            use_container_width=True, hide_index=True)

    _show_tech_errors(data)

    # --- Kirim ke Telegram (semua gaya) ---
    st.divider()
    msg = format_signals_message(data["per_style"])
    tcol1, tcol2 = st.columns([1, 3])
    if tcol1.button("📤 Kirim ke Telegram", key="send_tg"):
        token, recipients = notify_telegram.load_config()
        if not token:
            st.error("Token belum diset — isi TELEGRAM_BOT_TOKEN di .env "
                     "(atau config.json → telegram.token).")
        elif not recipients:
            st.error("Belum ada penerima. Tambah chat_id di panel kiri → 📤 Telegram.")
        else:
            results = notify_telegram.broadcast(msg, token, recipients)
            ok = sum(1 for _, o, _ in results if o)
            if ok:
                st.success(f"Terkirim ke {ok}/{len(results)} penerima.")
            for cid, o, desc in results:
                if not o:
                    hint = ""
                    if "chat not found" in desc.lower():
                        hint = (" → Penerima harus kirim **/start** ke bot dulu "
                                "(bot tak bisa memulai chat). Untuk grup, tambahkan bot sebagai anggota.")
                    st.warning(f"Gagal ke {cid}: {desc}{hint}")
    tcol2.caption("Mengirim sinyal BELI yang actionable ke semua penerima terdaftar.")
    with st.expander("👁️ Pratinjau pesan Telegram"):
        st.text(msg)

    render_journal()


def render_journal():
    st.divider()
    st.subheader("📊 Winrate & Jurnal Trade")
    open_pos = tracker.load_open()
    wk = tracker.winrate("W")
    mo = tracker.winrate("M")

    closed = sum(r["Trade"] for r in mo)
    wins = sum(r["Win"] for r in mo)
    m1, m2, m3, m4 = st.columns(4)
    m1.metric("Posisi terbuka", len(open_pos))
    m2.metric("Trade dihitung", closed)
    m3.metric("Win", wins)
    m4.metric("Winrate total", f"{(wins / closed * 100):.0f}%" if closed else "—")

    col_w, col_m = st.columns(2)
    with col_w:
        st.caption("**Per minggu**")
        st.dataframe(pd.DataFrame(wk) if wk else pd.DataFrame(
            columns=["Periode", "Trade", "Win", "Loss", "Winrate %"]),
            hide_index=True, use_container_width=True)
    with col_m:
        st.caption("**Per bulan**")
        st.dataframe(pd.DataFrame(mo) if mo else pd.DataFrame(
            columns=["Periode", "Trade", "Win", "Loss", "Winrate %"]),
            hide_index=True, use_container_width=True)

    if open_pos:
        # "Win berjalan": posisi terbuka yang sudah kena TP1 -> SL di breakeven,
        # tak bisa lagi jadi loss. Ditandai visual; belum masuk winrate realized.
        berjalan = sum(1 for p in open_pos.values() if p.get("hits"))
        if berjalan:
            st.caption(f"🟢 **{berjalan}** posisi sudah **Win berjalan** (kena TP1, SL "
                       "naik ke breakeven — tak bisa jadi loss) dan **sudah dihitung "
                       "Win** di statistik, walau posisi masih terbuka.")
        st.caption("**Posisi terbuka (dipantau)**")
        rows = [{"Gaya": p["style"], "Saham": p["ticker"], "Sinyal": p["signal"],
                 "Status": "🟢 Win berjalan" if p.get("hits") else "⏳ dipantau",
                 "Entry": p["entry_ref"], "SL": p["stop"],
                 "TP1": p.get("tp1"), "TP2": p.get("tp2"), "TP3": p.get("tp3"),
                 "TP kena": ", ".join(p.get("hits", [])) or "—"}
                for p in open_pos.values()]
        st.dataframe(pd.DataFrame(rows).style.format(
            {"Entry": "{:,.0f}", "SL": "{:,.0f}", "TP1": "{:,.0f}",
             "TP2": "{:,.0f}", "TP3": "{:,.0f}"}, na_rep="-"),
            hide_index=True, use_container_width=True)

    # --- Unduh jurnal (Excel) untuk analisa sendiri ---
    if os.path.exists(tracker.EVENTS_CSV):
        try:
            st.download_button(
                "⬇️ Download Jurnal (Excel .xlsx)", build_journal_xlsx(),
                file_name="jurnal_sinyal.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        except Exception as e:
            st.caption(f"Gagal membuat Excel: {e}")
    else:
        st.caption("Belum ada jurnal untuk diunduh.")
    st.caption(f"Jurnal tersimpan di `{tracker.EVENTS_CSV}` · WIN = kena TP1+ "
               "(langsung dihitung, SL naik ke breakeven) · LOSS = kena SL tanpa TP.")


def _style_sheet(ws, df):
    """Rapikan sheet dengan warna lembut & profesional: header biru muda + teks
    navy, baris selang-seling abu-biru tipis, freeze header & autofilter."""
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    header_fill = PatternFill("solid", fgColor="DDEBF7")   # biru sangat muda
    header_font = Font(bold=True, color="1F4E78")          # navy lembut
    band_fill = PatternFill("solid", fgColor="F5F8FC")     # abu-biru tipis (zebra)
    header_border = Border(bottom=Side(style="medium", color="9DB7D6"))

    for i, col in enumerate(df.columns, 1):
        cell = ws.cell(row=1, column=i)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = header_border
        sample = [str(col)] + [str(v) for v in df[col].astype(str).values[:300]]
        ws.column_dimensions[cell.column_letter].width = min(max(max(map(len, sample)) + 2, 10), 42)
        if pd.api.types.is_numeric_dtype(df[col]):
            for r in range(2, ws.max_row + 1):
                ws.cell(row=r, column=i).number_format = "#,##0.##"

    for r in range(2, ws.max_row + 1):        # zebra baris genap
        if r % 2 == 0:
            for c in range(1, ws.max_column + 1):
                ws.cell(row=r, column=c).fill = band_fill

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def build_journal_xlsx():
    """Workbook Excel rapi: event jurnal + winrate mingguan/bulanan + posisi terbuka."""
    buf = io.BytesIO()
    events = (pd.read_csv(tracker.EVENTS_CSV)
              if os.path.exists(tracker.EVENTS_CSV) else pd.DataFrame())
    if not events.empty:
        events = events.rename(columns={
            "timestamp": "Waktu", "date": "Tanggal", "style": "Gaya", "ticker": "Saham",
            "signal": "Sinyal", "event": "Event", "price": "Harga", "entry": "Entry",
            "stop": "Stop", "tp1": "TP1", "tp2": "TP2", "tp3": "TP3",
            "gain_pct": "Gain %", "trade_id": "Trade ID"})
    open_pos = tracker.load_open()
    open_df = pd.DataFrame([
        {"Gaya": p["style"], "Saham": p["ticker"], "Sinyal": p["signal"],
         "Entry": p["entry_ref"], "Stop": p["stop"], "TP1": p.get("tp1"),
         "TP2": p.get("tp2"), "TP3": p.get("tp3"), "TP kena": ", ".join(p.get("hits", [])),
         "Dibuka": p.get("opened_at")}
        for p in open_pos.values()])

    sheets = {
        "Jurnal Event": events,
        "Winrate Mingguan": pd.DataFrame(tracker.winrate("W")),
        "Winrate Bulanan": pd.DataFrame(tracker.winrate("M")),
        "Posisi Terbuka": open_df,
    }
    with pd.ExcelWriter(buf, engine="openpyxl") as xw:
        for name, df in sheets.items():
            d = df if not df.empty else pd.DataFrame({"Info": ["(kosong)"]})
            d.to_excel(xw, sheet_name=name, index=False)
            _style_sheet(xw.sheets[name], d)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# Tab 2 — quality-value screener
# ---------------------------------------------------------------------------
def render_quality_value():
    st.caption("Skor fundamental: Quality (mutu bisnis) × Value (kemurahan). "
               "Selain 'Selected', tabel hanya menampilkan yang terbaik (verdict Quality Value), "
               "urut dari QV tertinggi.")
    if st.button("💎 Scan Quality Value", type="primary", key="scan_qv"):
        res, errors = scan_map(scan_tickers, qvscreener.evaluate, "Menilai")
        rows = []
        for r in res.values():
            f = r["fund"]
            rows.append({
                "Saham": f.get("ticker"), "Harga": f.get("price"),
                "Quality": r["quality"], "Value": r["value"], "QV": r["qv"],
                "F-Score": f.get("f_score"),
                "P/E": f.get("pe"), "P/B": f.get("pb"), "Div%": f.get("dividend_yield"),
                "ROE%": f.get("roe"),
                "NP Growth 3Y%": f.get("net_profit_cagr_3y"),
                "Rev Growth 3Y%": f.get("revenue_cagr_3y"),
                "D/E%": f.get("debt_to_equity"),
                "Harga Wajar": f.get("fair_value"),
                "MoS%": f.get("margin_of_safety"),
                "Verdict": r["verdict"],
            })
        st.session_state["qv"] = {"rows": rows, "errors": errors, "universe": universe,
                                  "results": {r["fund"]["ticker"]: r for r in res.values()}}

    data = st.session_state.get("qv")
    if not data:
        st.info("Klik **Scan Quality Value** untuk mulai.")
        return
    if not data["rows"]:
        st.error("Tidak ada data fundamental yang berhasil diambil.")
        return

    df = pd.DataFrame(data["rows"]).sort_values("QV", ascending=False, na_position="last")
    is_selected = data["universe"] == "Selected"
    good = df[df["Verdict"] == "Quality Value - menarik"]

    if is_selected:
        visible = df
    else:
        # Indexes / ALL: only the best names, capped so the table stays readable.
        visible = good.head(QV_TOP_N)

    c1, c2, c3 = st.columns(3)
    c1.metric("Saham dinilai", len(df))
    c2.metric("💎 Quality Value", len(good))
    c3.metric("Gagal ambil data", len(data["errors"]))

    title = "Peringkat Quality-Value" if is_selected else \
            f"Saham Terbaik — Quality Value (top {min(QV_TOP_N, len(good))})"
    st.subheader(title)
    st.caption("👆 Klik satu baris untuk melihat profil perusahaan, verdict, & alasan skornya.")
    if visible.empty:
        st.warning("Tidak ada saham dengan verdict 'Quality Value' pada universe ini.")
    else:
        visible = visible.reset_index(drop=True)
        event = st.dataframe(
            visible.style.map(color_verdict, subset=["Verdict"]).format({
                "Harga": "{:,.0f}", "Quality": "{:.0f}", "Value": "{:.0f}", "QV": "{:.0f}",
                "F-Score": "{:.0f}", "P/E": "{:.1f}", "P/B": "{:.1f}", "Div%": "{:.1f}",
                "ROE%": "{:.1f}", "NP Growth 3Y%": "{:.1f}", "Rev Growth 3Y%": "{:.1f}",
                "D/E%": "{:.1f}", "Harga Wajar": "{:,.0f}", "MoS%": "{:+.1f}"},
                na_rep="-"),
            use_container_width=True, hide_index=True,
            on_select="rerun", selection_mode="single-row", key="qv_table")
        picked = selected_ticker(event, visible)
        if picked and picked in data.get("results", {}):
            st.divider()
            qv_detail(data["results"][picked])

    st.caption("Quality = ROE, ROA, margin, Net Profit Growth 3Y, Revenue Growth 3Y, Debt/Equity. "
               "Value = P/E, P/B, dividend yield, PEG. F-Score = Piotroski (0–9). "
               "MoS = margin of safety vs harga wajar Graham (+ = di bawah harga wajar).")

    if data["errors"]:
        with st.expander(f"⚠️ {len(data['errors'])} saham gagal diambil"):
            for t, e in data["errors"]:
                st.write(f"**{t}**: {e}")


@st.fragment(run_every=AUTO_CHECK_SEC)
def _technical_auto():
    render_technical()


def render_mean_reversion():
    """Daily mean reversion screener that follows the PRD rules."""
    st.caption(
        "Screener swing harian berbasis mean reversion PRD: Bollinger Band, RSI, "
        "MA20/50/200, ATR, volume, support/resistance, dan risk/reward."
    )
    c1, c2, c3, c4 = st.columns([1, 1, 1, 1.4])
    mode_label = c1.radio("Mode", ["Conservative", "Aggressive"], key="mr_mode")
    min_score = c2.selectbox("Skor minimum", [50, 65, 80], index=1, key="mr_min_score")
    show_no_trade = c3.checkbox("Tampilkan No Trade", value=False, key="mr_show_no_trade")
    run = c4.button("🔍 Scan mean reversion", type="primary", key="mr_scan")
    c4.caption(f"{len(scan_tickers)} saham · {universe} · timeframe daily")

    mode = (
        mean_reversion.MODE_CONSERVATIVE
        if mode_label == "Conservative"
        else mean_reversion.MODE_AGGRESSIVE
    )
    sig = (universe, tuple(scan_tickers), mode)
    data = st.session_state.get("mr_data")
    if data and data.get("sig") != sig:
        data = None
        st.session_state.pop("mr_data", None)

    if run:
        if not scan_tickers:
            st.warning("Universe kosong.")
            return
        results, errors = scan_map(
            scan_tickers,
            lambda tk: mean_reversion.screen_one(tk, mode=mode),
            "Scan Mean Reversion",
        )
        rows, recos, analyses = [], [], {}
        for a in results.values():
            row, reco = build_mr_row(a)
            rows.append(row)
            analyses[a.instrument] = a
            if reco:
                recos.append(reco)
        hit_events, new_opens = _journal_mean_reversion_scan(analyses, mode_label)
        if hit_events:
            st.toast(f"{len(hit_events)} level Mean Reversion tersentuh")
        if new_opens:
            st.toast(f"{len(new_opens)} setup Mean Reversion masuk jurnal")
        if auto_send:
            token, recipients = notify_telegram.load_config()
            if token and recipients:
                for ev in hit_events:
                    notify_telegram.broadcast(format_event_message(ev), token, recipients)
        data = {"rows": rows, "recos": recos, "analyses": analyses, "errors": errors, "sig": sig}
        st.session_state["mr_data"] = data

    data = st.session_state.get("mr_data")
    if not data:
        st.info("Pilih mode lalu klik **Scan mean reversion** untuk mulai.")
        st.caption("Universe diambil dari sidebar yang sama dengan tab lain, jadi watchlist lama tetap bisa dipakai.")
        render_mr_backtest(mode, mode_label)
        return
    if not data.get("rows"):
        st.error("Tidak ada data yang berhasil diambil.")
        if data.get("errors"):
            with st.expander(f"⚠️ {len(data['errors'])} saham gagal diambil"):
                for t, e in data["errors"]:
                    st.write(f"**{t}**: {e}")
        render_mr_backtest(mode, mode_label)
        st.caption("Aplikasi ini hanya analisis teknikal berbasis data historis dan bukan rekomendasi investasi final.")
        return

    df = pd.DataFrame(data["rows"])
    required_cols = {"_rank", "Skor", "Sinyal"}
    if not required_cols.issubset(df.columns):
        st.error("Hasil scan belum lengkap. Klik **Scan mean reversion** lagi untuk memuat ulang data.")
        st.session_state.pop("mr_data", None)
        render_mr_backtest(mode, mode_label)
        st.caption("Aplikasi ini hanya analisis teknikal berbasis data historis dan bukan rekomendasi investasi final.")
        return

    df = df.sort_values(["_rank", "Skor"], ascending=False)
    visible = df[df["Skor"] >= min_score]
    if not show_no_trade:
        visible = visible[visible["Sinyal"] != mean_reversion.SIGNAL_NO_TRADE]
    visible = visible.drop(columns=["_rank"]).reset_index(drop=True)
    all_rows = df.drop(columns=["_rank"]).reset_index(drop=True)

    c1, c2, c3, c4, c5 = st.columns(5)
    c1.metric("Saham discan", len(all_rows))
    c2.metric("BUY", int((all_rows["Sinyal"] == mean_reversion.SIGNAL_BUY).sum()))
    c3.metric("WATCH", int((all_rows["Sinyal"] == mean_reversion.SIGNAL_WATCH).sum()))
    c4.metric("Ditampilkan", len(visible))
    c5.metric("Gagal ambil data", len(data["errors"]))

    st.caption("Kategori status mengikuti PRD: Strong 80-100, Valid 65-79, Weak 50-64, No Trade <50.")
    csv_bytes = all_rows.to_csv(index=False).encode("utf-8-sig")
    st.download_button(
        "⬇️ Export hasil scan (.csv)",
        csv_bytes,
        file_name=f"mean_reversion_{mode}.csv",
        mime="text/csv",
        key="mr_export_csv",
    )

    if visible.empty:
        st.warning("Tidak ada saham yang lolos filter tampilan saat ini.")
    else:
        event = st.dataframe(
            visible.style.map(color_signal, subset=["Sinyal"]).map(color_mr_status, subset=["Status"]).format(
                {
                    "Harga": "{:,.0f}",
                    "Chg%": "{:+.1f}",
                    "RSI 14": "{:.1f}",
                    "Jarak ke MA20%": "{:+.1f}",
                    "Vol Ratio": "{:.2f}",
                    "Entry": "{:,.0f}",
                    "Stop Loss": "{:,.0f}",
                    "Target 1": "{:,.0f}",
                    "Target 2": "{:,.0f}",
                    "R:R": "{:.2f}",
                    "Support": "{:,.0f}",
                    "Resistance": "{:,.0f}",
                },
                na_rep="-",
            ),
            use_container_width=True,
            hide_index=True,
            on_select="rerun",
            selection_mode="single-row",
            key="mr_table",
        )
        picked = selected_ticker(event, visible)
        if picked and picked in data["analyses"]:
            st.divider()
            mean_reversion_detail(data["analyses"][picked])

    st.subheader("💡 Ukuran Posisi (lot)")
    st.caption(f"Modal Rp {capital:,.0f} · risiko {risk_pct}%/trade · lot {cfg.get('lot_size', 100)}")
    if not data["recos"]:
        st.write("Belum ada setup BUY yang memenuhi syarat entry + reward/risk untuk mode ini.")
    else:
        st.dataframe(
            pd.DataFrame(data["recos"]).style.map(color_signal, subset=["Sinyal"]).format(
                {
                    "Entry": "{:,.0f}",
                    "Stop Loss": "{:,.0f}",
                    "Target 1": "{:,.0f}",
                    "Target 2": "{:,.0f}",
                    "R:R": "{:.2f}",
                    "Modal (Rp)": "{:,.0f}",
                    "Risiko (Rp)": "{:,.0f}",
                },
                na_rep="-",
            ),
            use_container_width=True,
            hide_index=True,
        )

    if data["errors"]:
        with st.expander(f"⚠️ {len(data['errors'])} saham gagal diambil"):
            for t, e in data["errors"]:
                st.write(f"**{t}**: {e}")

    render_mr_backtest(mode, mode_label)
    st.caption("Aplikasi ini hanya analisis teknikal berbasis data historis dan bukan rekomendasi investasi final.")


with tab_tech:
    _technical_auto()          # selalu aktif: bangun tiap 30s, scan tiap interval
with tab_qv:
    render_quality_value()
with tab_mr:
    render_mean_reversion()
